"""
code by songdaoyuan@202412.25
获取Gitee企业版仓库列表, 获取每个仓库的成员统计信息
"""

import os
import json
import time
import math
import requests
from dotenv import load_dotenv, find_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment


"""
Step 0
使用.env文件管理token信息
"""
_ = load_dotenv(find_dotenv(".env"))

access_token = os.environ["access_token"]  # 管理员账号的用户授权码
enterprise_id = os.environ["enterprise_id"]  # 企业ID

req_headers = {
    # 按需添加
}

"""
Step 1
获取所有的的仓库列表, 以JSON格式存储在本地
API文档链接: https://gitee.com/api/v8/swagger#/getEnterpriseIdProjectsProjectIdMembersWithTeamMembers
"""


def getRepoList(access_token: str, enterprise_id: str) -> int:
    url = f"https://api.gitee.com/enterprises/{enterprise_id}/projects"  # API路径
    params: dict = {
        "access_token": access_token,
        "per_page": 100,  # 每页最多返回100个仓库
        "page": 1,  # 初始页码为1
    }

    all_repos = []  # 用于存储所有仓库信息

    # 第一次请求，获取 total_count 和第一页的数据
    response = requests.get(url=url, params=params, headers=req_headers)

    if response.status_code == 200:
        data = response.json()
        total_count = data.get("total_count", 0)  # 仓库总数
        if total_count == 0:
            print("没有仓库数据，退出程序")
            return 0

        # 将第一页的数据添加到仓库列表
        all_repos.extend(data.get("data", []))

        # 计算总页数
        per_page = params["per_page"]
        total_pages = math.ceil(total_count / per_page)  # 向上取整
        print(f"总仓库数：{total_count}，每页：{per_page}，总页数：{total_pages}")

        # 从第2页开始循环获取剩余数据
        for page in range(2, total_pages + 1):
            params["page"] = page
            response = requests.get(url=url, params=params, headers=req_headers)

            if response.status_code == 200:
                data = response.json()
                all_repos.extend(data.get("data", []))  # 将当前页的数据追加到列表
                print(f"已获取第 {page} 页的数据，当前累计仓库总数：{len(all_repos)}")
            else:
                print(
                    f"请求第 {page} 页失败，状态码：{response.status_code}，错误信息：{response.text}"
                )
                return 0xC0FFEE

        # 将所有仓库数据保存到本地文件
        with open("RepoList.payload", "w", encoding="UTF-8") as file:
            json.dump(all_repos, file, ensure_ascii=False, indent=4)
        print(f"所有仓库数据已保存到 RepoList.payload，总计仓库数量：{len(all_repos)}")
        return 0

    else:
        print(f"请求失败，状态码：{response.status_code}, 错误信息：{response.text}")
        return 0xC0FFEE


# getRepoList(access_token, enterprise_id)

"""
Step 2
遍历每个仓库, 获取仓库的人员列表
API文档链接: https://gitee.com/api/v8/swagger#/getEnterpriseIdProjectsProjectIdMembersWithTeamMembers
"""


def getMembersList(access_token: str, enterprise_id: str, payload_name: str) -> int:

    with open(payload_name, "r", encoding=("UTF-8")) as file:
        repoJson: list = json.load(file)

    # 初始化一个二维列表repoList, 用于存储提取的项目的id和name
    repoList = []
    # 初始化一个二维列表projList, 用于存储最终的包含项目name和members的结果
    projList = []

    for item in repoJson:
        # 提取 id、name
        id_value = item.get("id")
        name_value = item.get("name")
        repoList.append([id_value, name_value])

    # 输出整合的结果
    # print(repoList)

    for eachRepo in repoList:

        project_id = eachRepo[0]
        project_name = eachRepo[1]

        url = f"https://api.gitee.com/enterprises/{enterprise_id}/projects/{project_id}/members/with_team_members"

        params: dict = {
            # 查询参数写在这里
            "access_token": access_token,
            "project_id": project_id,
            "page": "1",
            "per_page": "100",
        }

        # 初始化临时用于存储members信息的列表
        tmpMembersList = []

        response = requests.get(url=url, params=params, headers=req_headers)
        time.sleep(0.5)

        if response.status_code == 200:
            membersJson: list = response.json()

            for item in membersJson.get("data", []):
                # 获取当前repo的members
                name_value = item.get("name")
                # 将提取的值组成一个子列表
                tmpMembersList.append(name_value)

            print(tmpMembersList)
        else:
            print(
                f"请求失败, 状态码：{response.status_code}, 错误信息：{response.text}"
            )

        projList.append([project_name, tmpMembersList])

    print(projList)

    # Step 3: 保存到 Excel
    saveToExcel(projList, "ProjectMembers.xlsx")


"""
Step 3
将项目和成员列表保存到 Excel 文件
"""


def saveToExcel(projList, filename):
    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects and Members"

    # 写入表头
    ws.append(["项目名称", "成员名单"])

    # 写入数据
    for project in projList:
        project_name = project[0]
        members = project[1]
        # 将成员列表用换行符拼接为字符串
        members_joined = "\n".join(members)
        # 写入到工作表
        ws.append([project_name, members_joined])

    # 设置单元格样式，使第二列文本换行
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # 保存到文件
    wb.save(filename)
    print(f"数据已保存到 {filename}")


# 调用函数
getMembersList(access_token, enterprise_id, "RepoList.payload")